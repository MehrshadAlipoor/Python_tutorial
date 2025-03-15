# ==============================================
# Automation with Python - Working with Excel
# ==============================================

# This tutorial covers automating Excel files using Python and the `openpyxl` library.

# ----------------------------------------------
# 1. Installing openpyxl
# ----------------------------------------------
# Install openpyxl via pip if not installed:
# pip install openpyxl

# ----------------------------------------------
# 2. Creating and Saving an Excel File
# ----------------------------------------------
from openpyxl import Workbook  # Import Workbook from openpyxl

wb = Workbook()  # Create a new workbook
ws = wb.active  # Select the active worksheet
ws.title = "Data"  # Rename the worksheet

# Add column headers
ws.append(["Name", "Age", "City"])

# Save the workbook
wb.save("sample.xlsx")
print("Excel file created successfully!")

# ----------------------------------------------
# 3. Reading Data from an Excel File
# ----------------------------------------------
from openpyxl import load_workbook  # Import load_workbook to open existing files

wb = load_workbook("sample.xlsx")  # Open the existing Excel file
ws = wb.active  # Select the active worksheet

# Loop through rows and print values
for row in ws.iter_rows(values_only=True):
    print(row)

# ----------------------------------------------
# 4. Writing Data to an Excel File
# ----------------------------------------------
ws.append(["Alice", 25, "New York"])  # Add a new row
ws.append(["Bob", 30, "London"])  # Add another row

wb.save("sample.xlsx")  # Save the changes
print("Data added successfully!")

# ----------------------------------------------
# 5. Modifying Specific Cells
# ----------------------------------------------
ws["B2"] = 28  # Update Alice's age
wb.save("sample.xlsx")
print("Cell updated successfully!")

# ----------------------------------------------
# 6. Deleting Rows and Columns
# ----------------------------------------------
ws.delete_rows(2)  # Delete the second row
ws.delete_cols(2)  # Delete the second column (Age)

wb.save("sample.xlsx")
print("Row and column deleted successfully!")

# ----------------------------------------------
# 7. Formatting Cells (Bold Text, Colors)
# ----------------------------------------------
from openpyxl.styles import Font, PatternFill  # Import styling tools

# Apply bold text and red color to the first cell
ws["A1"].font = Font(bold=True, color="FF0000")
ws["A1"].fill = PatternFill(start_color="FFFF00", fill_type="solid")  # Yellow background

wb.save("sample.xlsx")
print("Formatting applied successfully!")

# ----------------------------------------------
# 8. Creating a Chart in Excel
# ----------------------------------------------
from openpyxl.chart import BarChart, Reference  # Import chart classes

chart = BarChart()
values = Reference(ws, min_col=2, min_row=1, max_row=3)  # Select data range
chart.add_data(values, titles_from_data=True)

ws.add_chart(chart, "E5")  # Place the chart at E5

wb.save("sample.xlsx")
print("Chart added successfully!")

# ----------------------------------------------
# 9. Automating Excel File Processing
# ----------------------------------------------
import os  # Import os to work with file directories

folder_path = "excel_files/"  # Define the folder containing Excel files

# Loop through all Excel files in the folder
for file in os.listdir(folder_path):
    if file.endswith(".xlsx"):
        wb = load_workbook(os.path.join(folder_path, file))  # Open each file
        ws = wb.active  # Select the active worksheet

        print(f"Processing {file}...")
        for row in ws.iter_rows(values_only=True):
            print(row)  # Print row data

# ----------------------------------------------
# Summary:
# ----------------------------------------------
# - Install `openpyxl` using pip
# - Create, read, write, and modify Excel files
# - Delete rows/columns and format cells
# - Add charts and automate multiple Excel files
# - Useful for data automation, reporting, and processing

# ===